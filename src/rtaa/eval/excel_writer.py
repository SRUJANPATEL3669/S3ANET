from __future__ import annotations

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path


# Conditional formatting colors
_RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
_GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_CLEAN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def _auto_fit_columns(ws):
    """Auto-fit column widths based on cell content."""
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 4, 30)


def _apply_asr_formatting(ws, asr_col_idx, start_row, end_row):
    """Color-code ASR values: red (high ASR = strong attack), green (low)."""
    for row_idx in range(start_row, end_row + 1):
        cell = ws.cell(row=row_idx, column=asr_col_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            if cell.value >= 70:
                cell.fill = _RED_FILL
            elif cell.value >= 40:
                cell.fill = _YELLOW_FILL
            else:
                cell.fill = _GREEN_FILL


def write_benchmark_results(results: list[dict], output_path: str | Path):
    """
    results: list of dicts, each containing:
    dataset, model, attack, OA, Kappa, AA, class_accs (list),
    train_time, test_time, total_time, SAM, SID, phys_consistency, ASR

    attack="Clean" rows are treated as baseline (no attack) and highlighted.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    datasets = sorted(set(r["dataset"] for r in results))

    header = [
        "Model", "Attack", "OA (%)", "Kappa", "AA (%)",
        "Train Time (s)", "Test Time (s)", "Total Time (s)",
        "SAM (deg)", "SID", "Phys. Consistency (%)", "ASR (%)"
    ]

    for ds in datasets:
        ws = wb.create_sheet(title=ds)
        ds_results = [r for r in results if r["dataset"] == ds]

        if not ds_results:
            continue

        n_classes = len(ds_results[0].get("class_accs", []))
        class_headers = [f"Class {i+1} OA (%)" for i in range(n_classes)]

        full_header = header[:5] + class_headers + header[5:]
        ws.append(full_header)

        # Style header row
        for col_idx, cell in enumerate(ws[1], 1):
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Find ASR column index
        asr_col_idx = len(full_header)  # Last column

        for r in ds_results:
            row = [
                r.get("model", ""),
                r.get("attack", ""),
                round(r.get("OA", 0.0), 2),
                round(r.get("Kappa", 0.0), 4),
                round(r.get("AA", 0.0), 2)
            ]
            class_accs = r.get("class_accs", [0.0] * n_classes)
            row.extend([round(a, 2) for a in class_accs])
            row.extend([
                round(r.get("train_time", 0.0), 1),
                round(r.get("test_time", 0.0), 3),
                round(r.get("total_time", 0.0), 1),
                round(r.get("SAM", 0.0), 4),
                round(r.get("SID", 0.0), 6),
                round(r.get("phys_consistency", 0.0), 2),
                round(r.get("ASR", 0.0), 2)
            ])
            ws.append(row)

            # Highlight clean (no-attack) rows
            if r.get("attack", "") == "Clean":
                for cell in ws[ws.max_row]:
                    cell.fill = _CLEAN_FILL

        # Apply ASR conditional formatting
        _apply_asr_formatting(ws, asr_col_idx, 2, ws.max_row)

        # Auto-fit columns
        _auto_fit_columns(ws)

    # Add a summary sheet
    ws_sum = wb.create_sheet(title="Summary", index=0)
    sum_header = [
        "Dataset", "Model", "Attack",
        "OA (%)", "AA (%)", "Kappa", "ASR (%)",
        "SAM (deg)", "SID"
    ]
    ws_sum.append(sum_header)
    for col_idx, cell in enumerate(ws_sum[1], 1):
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for r in results:
        ws_sum.append([
            r.get("dataset", ""),
            r.get("model", ""),
            r.get("attack", ""),
            round(r.get("OA", 0.0), 2),
            round(r.get("AA", 0.0), 2),
            round(r.get("Kappa", 0.0), 4),
            round(r.get("ASR", 0.0), 2),
            round(r.get("SAM", 0.0), 4),
            round(r.get("SID", 0.0), 6),
        ])

    # ASR formatting on summary
    _apply_asr_formatting(ws_sum, 7, 2, ws_sum.max_row)
    _auto_fit_columns(ws_sum)

    wb.save(output_path)
